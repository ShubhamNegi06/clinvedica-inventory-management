"""
Excel export.

Reuses the same filtering logic as the list endpoint
(sample_service.list_samples with page_size effectively unbounded) so
"export what I'm currently viewing" behaves identically to what's on
screen, rather than drifting from it via a separately-written query.

The output deliberately mirrors the Clinvedica template's exact two-row
header format (section-title row, then field-label row) rather than a
single flat header row — this means an exported file can be edited and
re-uploaded through bulk ingestion directly, matching the workflow
described where samples are managed in Excel and periodically synced in.
"""
import io
import uuid
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy.orm import Session

from app.models.field_definition import FieldDefinition, SECTION_LABELS, SECTION_ORDER
from app.models.user import User
from app.services import sample_service
from app.services.sample_service import FieldFilter


def _ordered_field_defs(db: Session) -> List[FieldDefinition]:
    """Field order matches the canonical section ordering (Case Details ->
    ... -> Biomarker Characterization) so exports read the same way the
    add-sample form and the original template are laid out."""
    field_defs = (
        db.query(FieldDefinition)
        .filter(FieldDefinition.is_active.is_(True))
        .order_by(FieldDefinition.section, FieldDefinition.display_order)
        .all()
    )
    section_rank = {s: i for i, s in enumerate(SECTION_ORDER)}
    field_defs.sort(key=lambda fd: (section_rank.get(fd.section, 999), fd.display_order))
    return field_defs


def export_samples_to_excel(
    db: Session,
    current_user: User,
    *,
    site_id: Optional[uuid.UUID] = None,
    field_filters: Optional[List[FieldFilter]] = None,
    search: Optional[str] = None,
) -> io.BytesIO:
    # page_size capped high but bounded — protects against an accidental
    # export of an unbounded dataset locking up a worker process.
    items, _total = sample_service.list_samples(
        db, current_user, site_id=site_id, field_filters=field_filters, search=search, page=1, page_size=10000
    )

    field_defs = _ordered_field_defs(db)

    wb = Workbook()
    ws = wb.active
    ws.title = "Samples"

    # --- Row 1: section titles (only in the first column of each section,
    #     matching the template's merged-cell appearance) ---
    fixed_section = "Case Details"
    section_row = [fixed_section, "", ""]  # Site, Subject ID, Sample ID all fall under Case Details
    last_section = None
    for fd in field_defs:
        label = SECTION_LABELS.get(fd.section, fd.section) if fd.section != last_section else ""
        section_row.append(label)
        last_section = fd.section
    ws.append(section_row)

    # --- Row 2: field labels ---
    header_row = ["Site ID", "Subject ID", "Sample ID"] + [fd.field_label for fd in field_defs]
    ws.append(header_row)
    for cell in ws[1]:
        cell.font = Font(bold=True, italic=True)
    for cell in ws[2]:
        cell.font = Font(bold=True)

    # --- Data rows ---
    for sample in items:
        row = [str(sample.site_id), sample.subject_id, sample.sample_id]
        row += [sample.custom_fields.get(fd.field_key, "") for fd in field_defs]
        ws.append(row)

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
